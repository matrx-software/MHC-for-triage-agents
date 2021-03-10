# -*- coding: utf-8 -*-
"""
Created on Tue Jul 14 16:39:11 2020

@author: birgi
"""

import random
from openpyxl import load_workbook
import csv
import pandas as pd
import numpy as np
from mhc.sickness_model import SicknessModel, init_patient_specific_offsets

# gegevens in te voeren door gebruiker:
n_patients = 10  # number of patients that is generated
min_age = 14  # minimal age of generated patients
max_age = 95  # maximum age of generated patients
n_sims = 10  # number simulations per patient
random_seed = 10


# Simulate a patient's progression given some care
def simulate(sm, patient, medical_care):
    # TODO Some patient configuration seem to result in no improvement (or detoriation), hence for now a maximum number
    #  of 'ticks' the simulation runs.
    max_ticks = 10e4

    # First get the patients initial data
    symptoms = patient[7]
    symptoms_start = None  # TODO: Seems not to be used by the SicknessModel, so not used here either
    fitness = patient[6]
    patient_medical_offsets = init_patient_specific_offsets()

    old_health = None
    health = None
    n_tick = 0
    delta = 0
    while not np.isnan(n_tick) and (health is None or 0 < health < 100):
        n_tick += 1
        health, symptoms = sm.update_sickness(health,
                                              symptoms,
                                              symptoms_start,
                                              fitness,
                                              medical_care,
                                              patient_medical_offsets)

        delta += health - old_health if old_health is not None else 0
        old_health = health

        if n_tick > 10000 and n_tick % 10000 == 0:
            print(f"Simulation is taking a while: @{n_tick}/{max_ticks} ({np.round(n_tick / max_ticks * 100, 4)}%)")

        if n_tick >= max_ticks:
            n_tick = np.NaN

    # The faster the health change, the higher the uncertainty over the survical chance
    # (e.g.: the faster someone survives/dies suddenly)
    delta = delta / n_tick
    delta = delta / 3  # TODO Is this indeed the maximum change of health? max care change=2 and offset=0.0 +/- 0.8
    stdev = min(0.25, max(0, abs(delta * 0.25)))  # max stdev of 0.25

    if health >= 100:
        survival_chance = max(0, min(1.0, np.random.normal(0.8, stdev)))
    else:
        survival_chance = max(0, min(1.0, np.random.normal(0.2, stdev)))

    return survival_chance, n_tick

# make python list from female and male names in excel
wb = load_workbook("namen.xlsx")  # Work Book
ws = wb['Top_eerste_voornamen_NL_2010']  # Work Sheet
column = ws['A']  # Column
names_female = [column[x].value for x in range(len(column))]
column2 = ws['B']  # Column
names_male = [column2[x].value for x in range(len(column2))]

# set the random seed
random.seed(random_seed)

patient_data = []
genders = ["Man", "Vrouw"]
# define possible professions (besides "scholier" and "gepensioneerd")
professions = ["Software ontwikkeling manager", "Medisch specialist", "Piloot", "Software engineer", "Bedrijfsadviseur",
               "Advocaat", "Apotheker", "Tandarts", "Apotheek manager", "Arts", "Rechercheur", "Bedrijfsdirecteur",
               "Kapper", "Schoonmaker", "Postbode", "Kok", "Cassière", "Winkel verkoper", "Afwasser",
               "Wasserij personeel", "Glazenwasser", "Lopende band medewerker", "Kleermaker", "Vuilnisman"]
fit_options = ["Zeer laag", "Laag", "Gemiddeld", "Hoog", "Zeer hoog"]

# TODO These did not map to what the sickness model expected, hence change it to the next rule. As I could not find
#  anywhere where the translatation from this old set was made to the set the sicknessmodel expects
# symptoms = ["zeer hoog", "hoog", "gemiddeld", "laag", "zeer laag"]
symptoms = ["Zeer hoog", "Hoog", "Gemiddeld", "Mild", "Zeer mild"]

for n in range(n_patients):
    # generate gender, fitness level, symptom seriousness and age of patient
    gender = random.choice(genders)
    fit = random.choice(fit_options)
    symptom = random.choice(symptoms)
    age = random.randrange(min_age, max_age)

    # generate name according to gender
    if gender == "Man":
        name = random.choice(names_male)
    else:
        name = random.choice(names_female)
    # generate profession according to age
    if age < 16 or (age < 18 and random.random() < 0.5):
        profession = "Scholier"
        homeSituation = "Thuiswonend kind"
    elif age > 67:
        profession = "Gepensioneerde, beroep was " + random.choice(professions)
    else:
        profession = random.choice(professions)

    p = random.random()  # create random number
    n_children = random.randrange(1, 5)  # create random number between 1 and 5 for number of children
    if age < 16:
        homeSituation = "Thuiswonend kind"
    elif age < 18:
        if p < 0.7:
            homeSituation = "Thuiswonend kind"
        else:
            homeSituation = "Alleenwonend"
    elif age < 25:
        if p < 0.3:
            homeSituation = "Thuiswonend kind"
        elif p < 0.7:
            homeSituation = "Alleenwonend"
        elif p < 0.9:
            homeSituation = "Getrouwd"
        else:
            homeSituation = "Getrouwd met " + str(n_children) + " kind(eren)"
    elif age < 35:
        if p < 0.1:
            homeSituation = "Thuiswonend kind"
        elif p < 0.4:
            homeSituation = "Alleenwonend"
        elif p < 0.65:
            homeSituation = "Getrouwd"
        elif p < 0.9:
            homeSituation = "Getrouwd met " + str(n_children) + " kind(eren)"
        else:
            homeSituation = "Gescheiden"
    elif age < 50:
        if p < 0.15:
            homeSituation = "Alleenwonend"
        elif p < 0.35:
            homeSituation = "Getrouwd"
        elif p < 0.65:
            homeSituation = "Getrouwd met " + str(n_children) + " kind(eren)"
        elif p < 0.95:
            homeSituation = "Gescheiden"
        else:
            homeSituation = "Weduwe"
    elif age < 65:
        if p < 0.1:
            homeSituation = "Alleenwonend"
        elif p < 0.35:
            homeSituation = "Getrouwd"
        elif p < 0.65:
            homeSituation = "Getrouwd met " + str(n_children) + " kind(eren)"
        elif p < 0.9:
            homeSituation = "Gescheiden"
        else:
            homeSituation = "Weduwe"
    elif age < 80:
        if p < 0.1:
            homeSituation = "Alleenwonend"
        elif p < 0.3:
            homeSituation = "Getrouwd"
        elif p < 0.5:
            homeSituation = "Getrouwd met " + str(n_children) + " kind(eren)"
        elif p < 0.7:
            homeSituation = "Gescheiden"
        else:
            homeSituation = "Weduwe"
    else:
        if p < 0.05:
            homeSituation = "Alleenwonend"
        elif p < 0.1:
            homeSituation = "Getrouwd"
        elif p < 0.25:
            homeSituation = "Getrouwd met " + str(n_children) + " kind(eren)"
        elif p < 0.4:
            homeSituation = "Gescheiden"
        else:
            homeSituation = "Weduwe"
    patient_data.append([n, name, gender, age, profession, homeSituation, fit, symptom])
# print(patient_data[1:10])

#################################################################
# Generate Data for survival prediction based on sickness model #
#################################################################
sm = SicknessModel(config=None)  # TODO Does not seem to use the config, so not used
for idx, p in enumerate(patient_data):
    # Simulate patient progression based on each treatment
    survival_eerste_hulp = []
    duration_eerste_hulp = []
    survival_huis = []
    duration_huis = []
    survival_ziekenboeg = []
    duration_ziekenboeg = []
    survival_ic = []
    duration_ic = []

    for _ in range(n_sims):
        # No treatment
        survival, n_ticks = simulate(sm, p, medical_care="eerste hulp")
        survival_eerste_hulp.append(survival)
        duration_eerste_hulp.append(n_ticks)

        # Send to home
        survival, n_ticks = simulate(sm, p, medical_care="huis")
        survival_huis.append(survival)
        duration_huis.append(n_ticks)

        # Send to ward
        survival, n_ticks = simulate(sm, p, medical_care="ziekenboeg")
        survival_ziekenboeg.append(survival)
        duration_ziekenboeg.append(n_ticks)

        # Send to IC
        survival, n_ticks = simulate(sm, p, medical_care="IC")
        survival_ic.append(survival)
        duration_ic.append(n_ticks)

    # calculate means and deviations
    mean_survival_eerste_hulp = np.mean(survival_eerste_hulp)
    stdev_survival_eerste_hulp = np.std(survival_eerste_hulp)
    mean_duration_eerste_hulp = np.mean(duration_eerste_hulp)
    stdev_duration_eerste_hulp = np.std(duration_eerste_hulp)

    mean_survival_huis = np.mean(survival_huis)
    stdev_survival_huis = np.std(survival_huis)
    mean_duration_huis = np.mean(duration_huis)
    stdev_duration_huis = np.std(duration_huis)

    mean_survival_ziekenboeg = np.mean(survival_ziekenboeg)
    stdev_survival_ziekenboeg = np.std(survival_ziekenboeg)
    mean_duration_ziekenboeg = np.mean(duration_ziekenboeg)
    stdev_duration_ziekenboeg = np.std(duration_ziekenboeg)

    mean_survival_ic = np.mean(survival_ic)
    stdev_survival_ic = np.std(survival_ic)
    mean_duration_ic = np.mean(duration_ic)
    stdev_duration_ic = np.std(duration_ic)

    survival_chances = [mean_survival_eerste_hulp, stdev_survival_eerste_hulp,
                        mean_survival_huis, stdev_survival_huis,
                        mean_survival_ziekenboeg, stdev_survival_ziekenboeg,
                        mean_survival_ic, stdev_survival_ic]
    duration = [mean_duration_eerste_hulp, stdev_duration_eerste_hulp,
                mean_duration_huis, stdev_duration_huis,
                mean_duration_ziekenboeg, stdev_duration_ziekenboeg,
                mean_duration_ic, stdev_duration_ic]

    # Append
    patient_data[idx].extend(survival_chances)
    patient_data[idx].extend(duration)

    if idx % 10 == 0:
        print(f"@{idx}/{n_patients} ({np.round(idx/n_patients*100,4)}%)")

#############################################
# Generate remaining life years per patient #
#############################################
for idx, p in enumerate(patient_data):
    age = p[3]
    fitness_mapping = {"Zeer laag": -2, "Laag": -1, "Gemiddeld": 0, "Hoog": 1, "Zeer hoog": 2}
    symptom_mapping = {"Zeer mild": 1, "Mild": 3, "Gemiddeld": 5, "Hoog": 7, "Zeer hoog": 9}
    fitness = fitness_mapping[p[6]]
    symptoms = symptom_mapping[p[7]]
    factor = (fitness+5-symptoms)/20
    rnd_years = np.random.randint(-5, 6)
    delta = factor * (max_age - age)
    years = min(1, (delta+ rnd_years))
    patient_data[idx].extend([years])

####################################
# convert patient data to csv file #
####################################
patient_data = pd.DataFrame(patient_data, columns=["index", "name", "gender", "age", "profession", "home_situation",
                                                   "fitness", "symptoms",
                                                   "survival_eerste_hulp", "std_survival_eerste_hulp",
                                                   "survival_huis", "std_survival_huis",
                                                   "survival_ziekenboeg", "std_survival_ziekenboeg",
                                                   "survival_IC", "std_survival_IC",
                                                   "opnameduur_eerste_hulp", "std_opnameduur_eerste_hulp",
                                                   "opnameduur_huis", "std_opnameduur_huis",
                                                   "opnameduur_ziekenboeg", "std_opnameduur_ziekenboeg",
                                                   "opnameduur_IC", "std_opnameduur_IC",
                                                   "remaining_life_years"])
patient_data.to_csv("patient_data.csv", index=False, sep=";")

print("Written output to 'patient_data.csv'")
# thuiswonend kind/alleenwonend/ samenwonend/ getrouwd/ getrouwd met x kinderen/ gescheiden/ weduw(naar)e)
