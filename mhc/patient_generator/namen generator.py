# -*- coding: utf-8 -*-
"""
Created on Thu Jul 16 10:03:56 2020

@author: birgi
"""
from openpyxl import load_workbook
import random

wb = load_workbook("namen.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Top_eerste_voornamen_NL_2010')  # Work Sheet
column = ws['A']  # Column
names_female = [column[x].value for x in range(len(column))]
column2 = ws['B']  # Column
names_male = [column2[x].value for x in range(len(column2))]
#print(names_female)
#print(names_male)

n_children = random.randrange(1,5)
homeSituation = "getrouwd met " + str(n_children) + " kind(eren)"
print(homeSituation)